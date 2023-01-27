import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Create a list of domains to query
domains = ["example1.com", "example2.com", "example3.com"]

# Create a new workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Results"

# Add column headers to the worksheet
worksheet.cell(row=1, column=1).value = "Procurar"
worksheet.cell(row=1, column=2).value = "Situação"

# Create a webdriver instance
options = webdriver.ChromeOptions()
options.add_argument("--headless")
driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=options)

# Iterate through the list of domains
for i, domain in enumerate(domains, start=2):
    # Go to the registry.com website and search for the domain
    driver.get("https://registro.br/")
    search_box = driver.find_element(By.ID, "is-avail-field")
    search_box.send_keys(domain)
    search_box.submit()

    # Wait for the results to load
    wait = WebDriverWait(driver, 10)
    element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'strong[class="font-3 fullCost"]')))

    # Extract the data from the search results
    result = driver.find_element(By.CSS_SELECTOR, 'strong[class="font-3 fullCost"]')
    print("Results for", domain)
    print(result.text)

    # Write the domain and results to the worksheet
    worksheet.cell(row=i, column=1).value = domain
    worksheet.cell(row=i, column=2).value = result.text

# Save the workbook
workbook.save("results.xlsx")

# Close the webdriver
driver.quit()
